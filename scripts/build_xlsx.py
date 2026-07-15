import json, os, sys
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..")
papers = json.load(open(os.path.join(BASE, "data", "papers.json")))
TODAY = sys.argv[1] if len(sys.argv) > 1 else date.today().isoformat()


def link_for(p):
    return p.get("url") or ("https://doi.org/" + p["doi"] if p.get("doi") else "")

for p in papers:
    p["link"] = link_for(p)

# Journal display order (from data/journals.json — edit that file to add journals)
JOURNALS = json.load(open(os.path.join(BASE, "data", "journals.json")))
ORDER = [j["name"] for j in JOURNALS]
OV2 = json.load(open(os.path.join(BASE, "data", "overview.json")))
TLABEL = {t["id"]: t["label"] for t in OV2.get("themes_taxonomy", [])}
papers.sort(key=lambda p:(ORDER.index(p["journal"]) if p["journal"] in ORDER else 99, p.get("date",""))[::1])
papers.sort(key=lambda p:(ORDER.index(p["journal"]) if p["journal"] in ORDER else 99))

wb = Workbook()

# ---------- Palette ----------
NAVY   = "1F3864"
BLUE   = "2E5496"
LIGHT  = "D9E1F2"
LIGHT2 = "EAF0FA"
GREY   = "595959"
FONT = "Arial"

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# =========================================================
#  SHEET 1: Overview
# =========================================================
ov = wb.active
ov.title = "Overview"
ov.sheet_view.showGridLines = False

def setcell(ws, coord, val, *, size=11, bold=False, color="000000", fill=None,
            wrap=False, italic=False, align="left", valign="top"):
    c = ws[coord]
    c.value = val
    c.font = Font(name=FONT, size=size, bold=bold, color=color, italic=italic)
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    c.alignment = Alignment(wrap_text=wrap, horizontal=align, vertical=valign)
    return c

ov.column_dimensions["A"].width = 3
ov.column_dimensions["B"].width = 104

setcell(ov,"B2","Regional Science & Regional Economics — Research Briefing", size=20, bold=True, color=NAVY)
setcell(ov,"B3",f"Rolling archive across 15 journals · last updated {TODAY}", size=11, color=GREY, italic=True)

r = 5
def para(text, *, size=11, bold=False, color="000000", gap_before=0, fill=None, wrap=True, italic=False):
    global r
    r += gap_before
    ov.row_dimensions[r].height = None
    setcell(ov, f"B{r}", text, size=size, bold=bold, color=color, wrap=wrap, fill=fill, italic=italic)
    r += 1

def header(text):
    global r
    r += 1
    setcell(ov, f"B{r}", text, size=13, bold=True, color="FFFFFF", fill=BLUE, wrap=True)
    ov.row_dimensions[r].height = 22
    r += 1

# Snapshot
header("At a glance")
from collections import Counter
nj = len(set(p["journal"] for p in papers))
withabs = sum(1 for p in papers if p["abstract"])
para(f"This briefing indexes {len(papers)} recent items across {nj} journals (research articles plus a few "
     f"book reviews / corrigenda, flagged in the Type column). Full abstracts were captured for {withabs}; "
     f"the remainder — mostly Elsevier (ScienceDirect) and Taylor & Francis titles — are listed with complete "
     f"title/author/DOI metadata, and each abstract is one click away via the DOI link (use your UW–Madison "
     f"library access). Every paper lives on the 'All papers' tab, sortable and filterable by journal, date, and keyword.")

# Themes
header("What's going on — dominant themes this quarter")
themes = [
 ("1. China's regional & urban transformation dominates the empirical output.",
  "By volume the single largest cluster: subway construction and firm taxation, treaty-port legacies on inter-city "
  "connectivity, city–county merger reforms and the gender wage gap, urban innovation/patent networks, digital-village "
  "governance, farmland irrigation institutions, and agricultural 'digital intelligence.' China is now the default "
  "empirical laboratory for identification-driven regional economics."),
 ("2. Climate, green transition & the environment cut across every journal.",
  "Heat and traffic accidents (JRS), floods and urban density (JUE), the urban–rural income gap during the green "
  "transition, soil-carbon markets (EPA), circular economy in shrinking regions, environmental crime, and air-quality "
  "capitalization into housing. 'Green transition geography' has become a named subfield (see Papers in Regional Science)."),
 ("3. Housing, real estate & financialization is the hottest urban-econ topic.",
  "Rent control and housing-wealth redistribution (St Paul), house prices crowding out SME lending, Airbnb and eviction "
  "filings, immigration salience and housing wealth, plus critical-geography work on Beirut, Johannesburg and property-tax "
  "redistribution. Causal-ID and political-economy lenses sit side by side."),
 ("4. Place-based policy & regional inequality — the enduring core question.",
  "Italy's inner-areas strategy, EU structural funds, China's development zones, the Community Reinvestment Act, and a "
  "useful new mechanism-based framework for the *unintended* consequences of place- and person-based policy (Redinger & "
  "Ross). Scale-sensitive inequality measurement is back (Theil decomposition, IRSR)."),
 ("5. A transportation-economics moment at RSUE.",
  "Regional Science and Urban Economics is running what looks like a themed cluster of invited surveys — maritime shipping "
  "networks, interstate highways and rural America, transport in low- and middle-income countries, and pre-industrial "
  "transport history. A strong one-stop set of literature reviews for anyone teaching transport/spatial economics."),
 ("6. The geography of AI & the digital economy is emerging fast.",
  "AI adoption across Dutch municipalities, AI innovation/exposure/use across European regions, post-ChatGPT AI skill demand "
  "in UK vacancies, digital transformation in Greater Manchester, and crowdfunding-driven industrial diversification."),
 ("7. Migration, mobility & labor markets.",
  "Work-from-home lengthening commutes, counterurban migration and structural change, skilled inter-regional migration in "
  "Spain, the spatial disability-employment gap, and the 'stickiness' of where you go to college."),
 ("8. Rural studies leans qualitative — care, food, and agroecology.",
  "Rural ageing and an 'ethic of care,' older people's food insecurity, regenerative agriculture and livestock tensions, "
  "alternative-food consumers, and social-farming cooperatives. Methodologically distinct from the econometric core — "
  "worth scanning for framing and fieldwork if your work touches ag/food systems."),
]
for h,b in themes:
    para(h, bold=True, color=NAVY, gap_before=0)
    para(b)

# Deep read
header("What to deep-read — a shortlist (tilted toward ag/applied & regional methods)")
para("Ten picks that are either methodologically useful, unusually clean in identification, or closest to applied "
     "agricultural/rural/regional economics. Full details and links on the 'All papers' tab.", italic=True, color=GREY)
picks = [
 ("Unintended consequences of place-based and person-based policies: a mechanism-based framework — Redinger & Ross (Annals of Regional Science)",
  "A conceptual scaffold (displacement, agglomeration, market access, behavioral response, intergenerational spillovers) that's immediately useful for framing any place-based-policy evaluation."),
 ("Impacts of Community Reinvestment Act designation on employment with spatial spillovers — Rupasingha & Goetz (Annals of Regional Science)",
  "Synthetic DID, explicit metro>micro>rural gradient and spillovers; ERS/Penn State applied-econ pedigree — a template for rural place-based causal work."),
 ("Transport in low- and middle-income countries — Storeygard (Regional Science and Urban Economics)",
  "Authoritative survey (informal transit, BRT, congestion, rural roads, colonial legacy). The reference to cite for development-transport work."),
 ("The urban wage premium in historical perspective — Butts, Jaworski & Kitchens (RSUE)",
  "Long-run (1940–2010) decomposition controlling for sorting; a clean methods read on agglomeration wage premia."),
 ("Uneven Greening: urban–rural income gap during the green transition — He, Chen & Yu (Journal of Regional Science)",
  "Directly relevant to distributional questions in ag/rural economics; documents a U-shaped green-transition/inequality relationship in China."),
 ("Out of Sight? Revealing Creativity-Led Innovation in Rural Regions — Castaldi, Cortinovis & Tessarin (Economic Geography, Open Access)",
  "Uses trademarks (not just patents) to make rural innovation visible — a measurement lesson for anyone studying non-metro innovation."),
 ("Counterurban migration and regional structural change — Neyse & Eriksson (Regional Science Policy & Practice)",
  "Post-pandemic counterurbanization meets regional structural change — squarely in the rural-revitalization conversation."),
 ("The redistribution of housing wealth caused by rent control — Ahern & Giacoletti (Journal of Urban Economics)",
  "Crisp natural experiment (St Paul 2021); quantifies who gains/loses from rent control — a model of policy-evaluation design."),
 ("Scales of Inequality: a two-stage Theil decomposition, US 1970–2020 — Lee & Breau (International Regional Science Review)",
  "Shows how the geographic scale you choose drives your inequality story; a methods must-read for spatial inequality work."),
 ("Older people and food insecurity in rural places — Milbourne, Blake & Chesworth (Journal of Rural Studies, Open Access)",
  "Qualitative depth on rural food insecurity and ageing — useful framing if your applied work touches food access."),
]
for i,(t,why) in enumerate(picks,1):
    para(f"{i}. {t}", bold=True)
    para(f"     Why: {why}", color=GREY)

# Coverage table note
header("Coverage & data notes")
para("• Window: most recent ~3 months, including online-first / articles-in-press. High-volume journals "
     "(Journal of Rural Studies, Land Use Policy) were capped at ~12–15 most-recent items; those caps are noted here so "
     "nothing reads as exhaustive.\n"
     "• Abstracts fully captured (80 of 159) for: J. Regional Science, Annals of Regional Science, Int'l Regional Science "
     "Review, Environment & Planning A, J. Urban Economics, J. Economic Geography (most), Economic Geography (1 of 3), "
     "Land Use Policy, Growth and Change — via publisher pages, IDEAS/RePEc and the OUCI Crossref mirror.\n"
     "• Abstracts NOT openly mirrored anywhere (publisher blocks + no Crossref/RePEc deposit): Papers in Regional Science, "
     "Regional Science Policy & Practice, RSUE in-press items, Journal of Rural Studies, Regional Studies, Spatial Economic "
     "Analysis. Every row has a working DOI link — open via UW–Madison library access for the abstract.\n"
     "• Land Use Policy rows now carry verified DOIs, full author lists and keywords from IDEAS/RePEc.", wrap=True)
ov.row_dimensions[r-1].height = 150

# =========================================================
#  SHEET 2: All papers
# =========================================================
ws = wb.create_sheet("All papers")
ws.sheet_view.showGridLines = False
cols = [("Journal",22),("Group",20),("Date / Issue",15),("Added",12),("Title",58),
        ("Authors",34),("Type",12),("Themes",26),("Keywords",30),("Abstract",80),("DOI / Link",40)]
for i,(h,w) in enumerate(cols,1):
    c = ws.cell(row=1, column=i, value=h)
    c.font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
    c.border = border
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 26

def ptype(p):
    t = p["title"].lower()
    if t.startswith("corrigendum") or t.startswith("erratum"):
        return "Corrigendum"
    if t.startswith("book review") or t=="book review":
        return "Book review"
    return "Research"

row = 2
for idx,p in enumerate(papers):
    shade = LIGHT2 if idx%2 else "FFFFFF"
    themes_str = "; ".join(TLABEL.get(t, t) for t in p.get("themes", []))
    vals = [p["journal"], p["group"], p.get("date",""), p.get("added",""), p["title"], p["authors"],
            ptype(p), themes_str, p["keywords"], p["abstract"] or "— (abstract not auto-retrieved; open via link)", ""]
    for j,v in enumerate(vals,1):
        c = ws.cell(row=row, column=j, value=v)
        c.font = Font(name=FONT, size=10, color="000000",
                      italic=(j==10 and not p["abstract"]))
        c.fill = PatternFill("solid", fgColor=shade)
        c.alignment = Alignment(wrap_text=(j in (5,6,8,9,10)), vertical="top", horizontal="left")
        c.border = border
    # link cell
    lc = ws.cell(row=row, column=11)
    if p["link"]:
        lc.value = p["link"]
        lc.hyperlink = p["link"]
        lc.font = Font(name=FONT, size=10, color="0563C1", underline="single")
    else:
        lc.value = "—"
        lc.font = Font(name=FONT, size=10, color=GREY)
    lc.fill = PatternFill("solid", fgColor=shade)
    lc.alignment = Alignment(wrap_text=True, vertical="top")
    lc.border = border
    row += 1

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:K{row-1}"

# =========================================================
#  SHEET 3: By-journal counts
# =========================================================
sc = wb.create_sheet("Journal index")
sc.sheet_view.showGridLines = False
sc.column_dimensions["A"].width = 3
for i,w in zip("BCDE",[34,15,12,18]):
    sc.column_dimensions[i].width = w
setcell(sc,"B2","Journals covered", size=16, bold=True, color=NAVY)
hdrs=["Journal","Items","Abstracts","Group"]
for i,h in enumerate(hdrs):
    c=sc.cell(row=4,column=2+i,value=h)
    c.font=Font(name=FONT,size=11,bold=True,color="FFFFFF");c.fill=PatternFill("solid",fgColor=NAVY)
    c.border=border;c.alignment=Alignment(horizontal="left")
rr=5
for jn in ORDER:
    grp=next((p["group"] for p in papers if p["journal"]==jn),"")
    n=sum(1 for p in papers if p["journal"]==jn)
    na=sum(1 for p in papers if p["journal"]==jn and p["abstract"])
    for i,v in enumerate([jn,n,na,grp]):
        c=sc.cell(row=rr,column=2+i,value=v)
        c.font=Font(name=FONT,size=10);c.border=border
        c.fill=PatternFill("solid",fgColor=LIGHT2 if rr%2 else "FFFFFF")
        c.alignment=Alignment(horizontal="left" if i in(0,3) else "center",wrap_text=(i==0))
    rr+=1
c=sc.cell(row=rr,column=2,value="TOTAL");c.font=Font(name=FONT,size=10,bold=True)
c=sc.cell(row=rr,column=3,value=len(papers));c.font=Font(name=FONT,size=10,bold=True);c.alignment=Alignment(horizontal="center")
c=sc.cell(row=rr,column=4,value=sum(1 for p in papers if p["abstract"]));c.font=Font(name=FONT,size=10,bold=True);c.alignment=Alignment(horizontal="center")

wb.save(os.path.join(BASE, "briefing.xlsx"))
print("saved. papers:",len(papers),"abstracts:",sum(1 for p in papers if p['abstract']))
